#!/usr/bin/env python3
"""
Test Excel extraction before deployment
Run this locally to validate your Excel structure
"""

import sys
from pathlib import Path
import pandas as pd
from openpyxl import load_workbook

EXCEL_FILE = Path('data/nooks_q2_model.xlsx')


def test_file_exists():
    """Check if file exists"""
    if EXCEL_FILE.exists():
        print(f"✅ File exists: {EXCEL_FILE}")
        return True
    else:
        print(f"❌ File not found: {EXCEL_FILE}")
        return False


def test_sheets():
    """List all sheets and check for expected sheet"""
    try:
        wb = load_workbook(EXCEL_FILE, read_only=True)
        sheets = wb.sheetnames
        print(f"\n📋 Available sheets: {', '.join(sheets)}")
        
        if 'Scenarios' in sheets:
            print(f"✅ Found expected sheet: 'Scenarios'")
            return True
        else:
            print(f"⚠️  Expected sheet 'Scenarios' not found")
            print(f"💡 Update CONFIG['sheet_name'] in scripts/extract-data.py to match your sheet name")
            return False
    except Exception as e:
        print(f"❌ Error reading sheets: {e}")
        return False


def test_data_structure():
    """Validate data structure"""
    try:
        df = pd.read_excel(EXCEL_FILE, sheet_name='Scenarios', header=None)
        print(f"\n📊 Data shape: {df.shape[0]} rows × {df.shape[1]} columns")
        
        print(f"\n📝 First 10 rows (first column):")
        for idx in range(min(10, len(df))):
            val = str(df.iloc[idx, 0]).strip()
            if val and val != 'nan':
                print(f"  Row {idx}: {val}")
        
        # Check for scenario prefixes
        found_scenarios = []
        for prefix in ['Ops', 'Bridge', 'Full']:
            for idx, row in df.iterrows():
                first_val = str(row.iloc[0]).strip()
                if first_val.startswith(prefix):
                    found_scenarios.append(prefix)
                    break
        
        print(f"\n🔍 Found scenario prefixes: {', '.join(set(found_scenarios))}")
        
        if len(set(found_scenarios)) == 3:
            print(f"✅ All 3 scenarios found")
            return True
        else:
            missing = set(['Ops', 'Bridge', 'Full']) - set(found_scenarios)
            print(f"⚠️  Missing scenarios: {', '.join(missing)}")
            return False
            
    except Exception as e:
        print(f"❌ Error reading data: {e}")
        return False


def test_extraction():
    """Try running the actual extraction"""
    try:
        print(f"\n🧪 Testing full extraction...")
        import subprocess
        result = subprocess.run(
            ['python', 'scripts/extract-data.py'],
            capture_output=True,
            text=True
        )
        
        if result.returncode == 0:
            print(f"✅ Extraction successful")
            print(result.stdout)
            return True
        else:
            print(f"❌ Extraction failed")
            print(result.stderr)
            return False
    except Exception as e:
        print(f"❌ Error running extraction: {e}")
        return False


def main():
    """Run all tests"""
    print("=" * 60)
    print("Excel Extraction Test Suite")
    print("=" * 60)
    
    tests = [
        ("File exists", test_file_exists),
        ("Sheet structure", test_sheets),
        ("Data structure", test_data_structure),
        ("Full extraction", test_extraction),
    ]
    
    results = []
    for name, test_func in tests:
        print(f"\n{'=' * 60}")
        print(f"Test: {name}")
        print(f"{'=' * 60}")
        result = test_func()
        results.append((name, result))
    
    # Summary
    print(f"\n{'=' * 60}")
    print("Summary")
    print(f"{'=' * 60}")
    
    for name, result in results:
        status = "✅" if result else "❌"
        print(f"{status} {name}")
    
    all_passed = all(r[1] for r in results)
    
    if all_passed:
        print(f"\n🎉 All tests passed! Ready to deploy.")
        return 0
    else:
        print(f"\n⚠️  Some tests failed. Fix issues before deploying.")
        return 1


if __name__ == '__main__':
    sys.exit(main())
