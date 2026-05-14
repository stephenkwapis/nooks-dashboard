#!/usr/bin/env python3
"""
Extract financial data from Nooks Excel model and generate JSON for dashboard
"""
import openpyxl
import json
from datetime import datetime
import sys

def safe_value(cell_value):
    """Convert cell value to JSON-safe format"""
    if cell_value is None:
        return None
    if isinstance(cell_value, datetime):
        return cell_value.strftime('%Y-%m-%d')
    if isinstance(cell_value, (int, float)):
        return float(cell_value) if not isinstance(cell_value, int) else int(cell_value)
    return str(cell_value)

def extract_data(excel_path):
    """Extract all dashboard data from Excel file"""
    wb = openpyxl.load_workbook(excel_path, data_only=True)
    data = {}
    
    # === RPT VISUALS SHEET ===
    if 'RPT Visuals' in wb.sheetnames:
        sheet = wb['RPT Visuals']
        
        # Quarterly data (row 21 columns C-F for Q1-Q4 revenue)
        # Row 25-35 for quarterly EBITDA (need to find exact row)
        data['quarterly'] = {
            'revenue': {
                'q1': safe_value(sheet['C21'].value),
                'q2': safe_value(sheet['D21'].value),
                'q3': safe_value(sheet['E21'].value),
                'q4': safe_value(sheet['F21'].value),
            }
        }
        
        # Annual projections (row 5 for revenue, row 15 for EBITDA, columns C-G for FY25-29)
        data['annual'] = {
            'revenue': [safe_value(sheet[f'{col}5'].value) for col in ['C', 'D', 'E', 'F', 'G']],
            'ebitda': [safe_value(sheet[f'{col}15'].value) for col in ['C', 'D', 'E', 'F', 'G']],
            'years': ['FY25A', 'FY26E', 'FY27E', 'FY28E', 'FY29E']
        }
    
    # === SCORECARD SHEET ===
    if 'Scorecard' in wb.sheetnames:
        sheet = wb['Scorecard']
        
        # Q1 Actuals vs AOP (row 6)
        data['scorecard'] = {
            'q1': {
                'revenue_actual': safe_value(sheet['C6'].value),
                'revenue_aop': safe_value(sheet['E6'].value),
                'ebitda_actual': safe_value(sheet['C12'].value),
                'ebitda_aop': safe_value(sheet['E12'].value),
            },
            'fy26': {
                'revenue_estimate': safe_value(sheet['I6'].value),
                'revenue_aop': safe_value(sheet['K6'].value),
                'ebitda_estimate': safe_value(sheet['I12'].value),
                'ebitda_aop': safe_value(sheet['K12'].value),
            }
        }
    
    # === BURN SHEET ===
    # NOTE: Cash balance data location needs to be verified in your Excel file
    # Current HTML has: [10265, 6563, 3403, 4940, 4679, 4786, 4789, 6882, 2426, 3156, 2888, 1448]
    # These are in thousands ($000s)
    # Please update the row number below to match your Excel structure
    
    data['cash_flow'] = {
        'months': ['Jan', 'Feb', 'Mar', 'Apr', 'May', 'Jun', 'Jul', 'Aug', 'Sep', 'Oct', 'Nov', 'Dec'],
        'cash_balance': [10265, 6563, 3403, 4940, 4679, 4786, 4789, 6882, 2426, 3156, 2888, 1448],  # Placeholder - update with Excel refs
        'net_changes': [3724, -3833, -2831, 1537, -261, 107, 2, 2093, -4456, 730, -268, -1441]  # Placeholder
    }
    
    # === BENCHMARKING SHEET ===
    if 'Benchmarking' in wb.sheetnames:
        sheet = wb['Benchmarking']
        
        # Revenue growth (rows 5-6, cols D-L for FY23-31)
        revenue_years = [safe_value(sheet[f'{col}5'].value) for col in ['D', 'E', 'F', 'G', 'H', 'I', 'J', 'K', 'L']]
        sqft_data = [safe_value(sheet[f'{col}7'].value) for col in ['F', 'G', 'H', 'I', 'J', 'K']]
        rev_per_sqft = [safe_value(sheet[f'{col}8'].value) for col in ['F', 'G', 'H', 'I', 'J', 'K']]
        
        data['benchmarking'] = {
            'revenue_growth': {
                'years': ['FY23A', 'FY24A', 'FY25A', 'FY26E', 'FY27E', 'FY28E', 'FY29E', 'FY30E', 'FY31E'],
                'revenue': revenue_years
            },
            'revenue_per_sqft': {
                'years': ['FY25A', 'FY26E', 'FY27E', 'FY28E', 'FY29E', 'FY30E'],
                'rev_per_sqft': rev_per_sqft,
                'op_sqft': sqft_data
            }
        }
    
    # === ADDITIONAL DASHBOARD DATA ===
    # NOTE: Add Excel cell references for these sections
    data['kpi_summary'] = {
        'site_revenue': {
            'fy26e': [11766, 4370, 6106, 19458],  # NL, COS, ES, Corp
            'aop': [13695, 5926, 8722, 19027],
            'sites': ['NL', 'COS', 'ES', 'Corp']
        }
    }
    
    data['corp_revenue'] = {
        'labels': ['ISE Contract', 'DARPA Prospero', 'TurboFCL', 'MSS Fees', 'Speed & Shield', 'Other'],
        'values': [8522, 3000, 1393, 1257, 940, 346]  # Update with Excel refs
    }
    
    data['headcount'] = {
        'by_team': {
            'teams': ['ISE Ops', 'Product', 'Corp IT', 'Growth', 'ISE Lead', 'Finance', 'NL Sec', 'Executive', 'NL IT', 'COS IT', 'HR', 'NL Ops', 'COS Sec', 'ES Sec', 'COS Ops', 'Corp Sec', 'MSS', 'ES Ops', 'ES IT', 'Legal'],
            'ftes': [24, 16, 15, 10, 7, 7, 6, 5, 5, 5, 5, 4, 4, 4, 4, 4, 3, 3, 3, 2]
        },
        'payroll_trend': {
            'quarters': ['Q1E', 'Q2E', 'Q3E', 'Q4E'],
            'actual': [3314, 4373, 5565, 5593],
            'aop': [3893, 4100, 4350, 4360]
        }
    }
    
    data['capex'] = {
        'retrofit': {
            'sites': ['National Landing', 'Colorado Springs', 'El Segundo', 'Corporate'],
            'fy26e': [10014, 761, 2727, 138],
            'aop': [9375, 246, 1235, 57]
        },
        'it_hardware': {
            'sites': ['National Landing', 'Colorado Springs', 'El Segundo', 'Corporate'],
            'fy26e': [2260, 620, 1367, 4495],
            'aop': [659, 645, 1365, 4934]
        }
    }
    
    # Add metadata
    data['_metadata'] = {
        'extracted_at': datetime.now().isoformat(),
        'source_file': excel_path.split('/')[-1]
    }
    
    return data

def main():
    if len(sys.argv) < 2:
        print("Usage: python extract_data.py <path_to_excel_file> [output_json_path]")
        sys.exit(1)
    
    excel_path = sys.argv[1]
    output_path = sys.argv[2] if len(sys.argv) > 2 else 'data/financial_data.json'
    
    print(f"Extracting data from: {excel_path}")
    data = extract_data(excel_path)
    
    # Create output directory if needed
    import os
    os.makedirs(os.path.dirname(output_path) if os.path.dirname(output_path) else '.', exist_ok=True)
    
    # Write JSON
    with open(output_path, 'w') as f:
        json.dump(data, f, indent=2)
    
    print(f"✓ Data extracted successfully to: {output_path}")
    print(f"  - Quarterly data: {len(data.get('quarterly', {}))} sections")
    print(f"  - Annual projections: {len(data.get('annual', {}).get('revenue', []))} years")
    print(f"  - Cash flow: {len(data.get('cash_flow', {}).get('months', []))} months")
    print(f"  - Benchmarking: {len(data.get('benchmarking', {}).get('revenue_growth', {}).get('years', []))} years")

if __name__ == '__main__':
    main()
