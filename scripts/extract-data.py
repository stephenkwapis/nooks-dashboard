#!/usr/bin/env python3
"""
Extract scenario data from Excel and output as JSON
"""

import sys
import json
from pathlib import Path
import pandas as pd
from openpyxl import load_workbook

# Configuration
EXCEL_FILE = Path('data/nooks_q2_model.xlsx')
OUTPUT_JSON = Path('data/scenarios.json')

# Excel structure configuration (adjust these based on your actual file)
CONFIG = {
    'sheet_name': 'Scenarios',  # Sheet containing the data
    
    # Define how to find each scenario's data
    # These are flexible - script will search for row labels
    'scenarios': {
        'ops': {
            'name': 'Ops / Base Case',
            'row_prefix': 'Ops',  # Looks for rows starting with "Ops"
        },
        'bridge': {
            'name': 'Series B — Bridge',
            'row_prefix': 'Bridge',
        },
        'full': {
            'name': 'Series B — Full',
            'row_prefix': 'Full',
        }
    },
    
    # Data columns to extract
    'metrics': ['Revenue', 'EBITDA', 'Payroll', 'Margin'],
    
    # Year columns (adjust based on your Excel structure)
    'year_columns': ['FY25A', 'FY26E', 'FY27E', 'FY28E', 'FY29E'],
}


def find_row_by_prefix(df, prefix):
    """Find row index where first column starts with prefix"""
    for idx, row in df.iterrows():
        first_val = str(row.iloc[0]).strip()
        if first_val.startswith(prefix):
            return idx
    return None


def extract_scenario_data(df, scenario_config):
    """Extract data for a single scenario"""
    
    prefix = scenario_config['row_prefix']
    data = {
        'name': scenario_config['name'],
        'revenue': [],
        'ebitda': [],
        'payroll': [],
        'margin': [],
    }
    
    # Find rows for each metric
    for metric in CONFIG['metrics']:
        row_label = f"{prefix} - {metric}"
        
        # Search for row
        for idx, row in df.iterrows():
            first_val = str(row.iloc[0]).strip()
            if first_val == row_label or first_val.startswith(f"{prefix}") and metric.lower() in first_val.lower():
                # Extract year values (assuming they're in columns 1-5 after the label)
                values = []
                for i in range(1, 6):  # 5 years
                    try:
                        val = row.iloc[i]
                        # Handle margin as percentage
                        if metric == 'Margin' and isinstance(val, (int, float)):
                            values.append(round(val * 100, 1) if val < 1 else round(val, 1))
                        else:
                            values.append(int(val) if isinstance(val, (int, float)) else 0)
                    except (IndexError, ValueError):
                        values.append(0)
                
                # Store in appropriate key
                key = metric.lower()
                data[key] = values
                break
    
    return data


def extract_from_excel():
    """Main extraction logic"""
    
    if not EXCEL_FILE.exists():
        print(f"❌ Excel file not found: {EXCEL_FILE}")
        sys.exit(1)
    
    print(f"📖 Reading Excel: {EXCEL_FILE}")
    
    try:
        # Try to read the configured sheet
        df = pd.read_excel(EXCEL_FILE, sheet_name=CONFIG['sheet_name'], header=None)
        print(f"✅ Found sheet: {CONFIG['sheet_name']}")
        
    except ValueError:
        # Sheet not found - list available sheets
        wb = load_workbook(EXCEL_FILE, read_only=True)
        print(f"❌ Sheet '{CONFIG['sheet_name']}' not found")
        print(f"📋 Available sheets: {wb.sheetnames}")
        sys.exit(1)
    
    # Extract data for each scenario
    scenarios = {}
    
    for key, config in CONFIG['scenarios'].items():
        print(f"🔍 Extracting '{config['name']}'...")
        scenario_data = extract_scenario_data(df, config)
        
        # Validate data
        if not scenario_data['revenue'] or sum(scenario_data['revenue']) == 0:
            print(f"⚠️  WARNING: No revenue data found for '{key}'")
        
        scenarios[key] = scenario_data
    
    return scenarios


def format_output(scenarios):
    """Format extracted data to match HTML structure"""
    
    output = {}
    
    for key, data in scenarios.items():
        # Calculate metrics
        revenue = data['revenue']
        ebitda = data['ebitda']
        
        # FY26 and FY29 revenue for KPIs
        rev26 = f"${revenue[1] / 1000:.1f}M" if len(revenue) > 1 else "—"
        rev29 = f"${revenue[4] / 1000:.0f}M" if len(revenue) > 4 else "—"
        
        # EBITDA margin for FY29
        margin = f"{data['margin'][4]}%" if len(data['margin']) > 4 else "—"
        
        # Calculate CAGR (FY25 to FY29)
        cagr = 0
        if len(revenue) == 5 and revenue[0] > 0 and revenue[4] > 0:
            cagr = (((revenue[4] / revenue[0]) ** (1/4)) - 1) * 100
        
        output[key] = {
            'name': data['name'],
            'rev26': rev26,
            'rev29': rev29,
            'margin': margin,
            'cagr': f"{cagr:.1f}%",
            'revenue': revenue,
            'ebitda': ebitda,
            'payroll': data['payroll'],
            'margin_array': data['margin'],
        }
    
    return output


def main():
    """Main execution"""
    
    print("=" * 60)
    print("Excel Data Extraction")
    print("=" * 60)
    
    # Extract raw data
    scenarios = extract_from_excel()
    
    # Format for output
    formatted = format_output(scenarios)
    
    # Save to JSON
    OUTPUT_JSON.parent.mkdir(parents=True, exist_ok=True)
    with open(OUTPUT_JSON, 'w') as f:
        json.dump(formatted, f, indent=2)
    
    print(f"\n✅ Data extracted to: {OUTPUT_JSON}")
    
    # Print summary
    print("\n📊 Summary:")
    for key, data in formatted.items():
        print(f"  {data['name']}:")
        print(f"    FY26E Revenue: {data['rev26']}")
        print(f"    FY29E Revenue: {data['rev29']}")
        print(f"    CAGR: {data['cagr']}")
    
    return 0


if __name__ == '__main__':
    sys.exit(main())
